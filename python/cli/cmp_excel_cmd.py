#!/usr/bin/env python3
"""cmp-excel 命令实现 - 对比两个 Excel 文件的 Ratio 值"""

import sys
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Optional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def parse_ratio(ratio_str: str) -> Optional[float]:
    """解析 Ratio 字符串为浮点数

    Args:
        ratio_str: Ratio 值，如 "1.14", "~1.14", "n/a" 等

    Returns:
        浮点数值，无效值返回 None
    """
    if not ratio_str or ratio_str.lower() == 'n/a':
        return None

    # 移除 ~ 前缀（表示近似值）
    ratio_str = ratio_str.strip().lstrip('~')

    try:
        return float(ratio_str)
    except (ValueError, TypeError):
        return None


def read_excel_ratios(file_path: str) -> Dict[str, Tuple[str, str]]:
    """读取 Excel 文件中的 Ratio 数据

    Args:
        file_path: Excel 文件路径

    Returns:
        Dict[benchmark_name, (ratio_str, ratio_float_str)]
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # 查找表头位置
        header_row = None
        ratio_col = None
        benchmark_col = None

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value == 'Ratio':
                    ratio_col = col_idx
                elif cell.value == 'Benchmark (Parameter)':
                    benchmark_col = col_idx

            if ratio_col and benchmark_col:
                header_row = row_idx
                break

        if not ratio_col or not benchmark_col:
            print(f"错误: 无法在 {file_path} 中找到 Ratio 或 Benchmark (Parameter) 列", file=sys.stderr)
            return {}

        # 读取数据
        result = {}
        for row in ws.iter_rows(min_row=header_row + 1):
            benchmark = row[benchmark_col - 1].value
            ratio = row[ratio_col - 1].value

            if benchmark and ratio:
                benchmark = str(benchmark).strip()
                ratio_str = str(ratio).strip()
                # 跳过分隔行
                if benchmark.startswith('-') or ratio_str.startswith('-'):
                    continue
                result[benchmark] = ratio_str

        wb.close()
        return result

    except Exception as e:
        print(f"错误: 读取 {file_path} 失败: {e}", file=sys.stderr)
        return {}


def print_results_table(results: List[dict], verbose: bool = False) -> None:
    """打印对比结果表格

    Args:
        results: 对比结果列表
        verbose: 是否详细输出（打印表格）
    """
    if not verbose or not results:
        return

    # ANSI 颜色代码
    GREEN = '\033[92m'    # 性能提升
    RED = '\033[91m'      # 性能劣化
    GRAY = '\033[90m'     # n/a
    BOLD = '\033[1m'
    RESET = '\033[0m'

    print(f"\n{BOLD}对比结果表格:{RESET}")
    print(f"{BOLD}{'Benchmark (Parameter)':<80} {'Before':>10} {'After':>10} {'Diff':>8}{RESET}")
    print("-" * 110)

    for result in results:
        benchmark = result['benchmark']
        if len(benchmark) > 77:
            benchmark = benchmark[:77] + "..."

        before = result['before']
        after = result['after']
        diff = result['diff']

        # 根据差异着色
        if result['matched'] and diff != 'n/a':
            try:
                diff_val = float(diff)
                if diff_val > 0:
                    diff_colored = f"{GREEN}{diff:>8}{RESET}"
                elif diff_val < 0:
                    diff_colored = f"{RED}{diff:>8}{RESET}"
                else:
                    diff_colored = f"{diff:>8}"
                print(f"{benchmark:<80} {before:>10} {after:>10} {diff_colored}")
            except ValueError:
                print(f"{benchmark:<80} {before:>10} {after:>10} {GRAY}{diff:>8}{RESET}")
        else:
            print(f"{GRAY}{benchmark:<80} {before:>10} {after:>10} {diff:>8}{RESET}")


def compare_ratios(before_file: str, after_file: str, output_file: Optional[str] = None, verbose: bool = False) -> bool:
    """对比两个 Excel 文件的 Ratio 值

    Args:
        before_file: Before Excel 文件路径
        after_file: After Excel 文件路径
        output_file: 输出文件路径（可选）
        verbose: 是否详细输出（打印表格）

    Returns:
        是否成功
    """
    print(f"读取 Before 文件: {before_file}")
    before_data = read_excel_ratios(before_file)
    if not before_data:
        print("错误: Before 文件无有效数据", file=sys.stderr)
        return False

    print(f"读取 After 文件: {after_file}")
    after_data = read_excel_ratios(after_file)
    if not after_data:
        print("错误: After 文件无有效数据", file=sys.stderr)
        return False

    print(f"\nBefore 文件共 {len(before_data)} 条记录")
    print(f"After 文件共 {len(after_data)} 条记录")

    # 合并所有 benchmark 名称
    all_benchmarks = set(before_data.keys()) | set(after_data.keys())

    # 分为匹配和不匹配的
    matched = []
    unmatched_before = []
    unmatched_after = []

    for benchmark in sorted(all_benchmarks):
        before_ratio = before_data.get(benchmark)
        after_ratio = after_data.get(benchmark)

        if before_ratio and after_ratio:
            matched.append(benchmark)
        elif before_ratio and not after_ratio:
            unmatched_before.append(benchmark)
        elif not before_ratio and after_ratio:
            unmatched_after.append(benchmark)

    print(f"匹配的 benchmark: {len(matched)}")
    print(f"仅在 Before 中: {len(unmatched_before)}")
    print(f"仅在 After 中: {len(unmatched_after)}")

    # 创建对比结果
    results = []

    # 处理匹配的 benchmark
    for benchmark in matched:
        before_ratio_str = before_data[benchmark]
        after_ratio_str = after_data[benchmark]

        before_ratio = parse_ratio(before_ratio_str)
        after_ratio = parse_ratio(after_ratio_str)

        if before_ratio is not None and after_ratio is not None:
            diff = before_ratio - after_ratio
            diff_str = f"{diff:+.2f}"
        else:
            diff_str = "n/a"

        results.append({
            'benchmark': benchmark,
            'before': before_ratio_str,
            'after': after_ratio_str,
            'diff': diff_str,
            'matched': True
        })

    # 处理不匹配的 benchmark (Before only)
    for benchmark in sorted(unmatched_before):
        results.append({
            'benchmark': benchmark,
            'before': before_data[benchmark],
            'after': 'n/a',
            'diff': 'n/a',
            'matched': False
        })

    # 处理不匹配的 benchmark (After only)
    for benchmark in sorted(unmatched_after):
        results.append({
            'benchmark': benchmark,
            'before': 'n/a',
            'after': after_data[benchmark],
            'diff': 'n/a',
            'matched': False
        })

    # 生成输出
    if not output_file:
        before_path = Path(before_file)
        after_path = Path(after_file)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"ratio_compare_{timestamp}.xlsx"

    # 创建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ratio Compare"

    # 样式
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 性能提升（绿色）、劣化（红色）、无变化（黄色）
    improve_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    degrade_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    na_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # 写入表头
    headers = ['Benchmark (Parameter)', 'Before Ratio', 'After Ratio', 'Ratio Diff']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 写入数据
    for row_idx, result in enumerate(results, 2):
        # Benchmark
        cell = ws.cell(row=row_idx, column=1, value=result['benchmark'])
        cell.alignment = left_align
        cell.border = thin_border

        # Before Ratio
        cell = ws.cell(row=row_idx, column=2, value=result['before'])
        cell.alignment = center_align
        cell.border = thin_border

        # After Ratio
        cell = ws.cell(row=row_idx, column=3, value=result['after'])
        cell.alignment = center_align
        cell.border = thin_border

        # Ratio Diff
        cell = ws.cell(row=row_idx, column=4, value=result['diff'])
        cell.alignment = center_align
        cell.border = thin_border

        # 根据差异着色
        if result['matched'] and result['diff'] != 'n/a':
            try:
                diff_val = float(result['diff'])
                if diff_val > 0:
                    cell.fill = improve_fill
                elif diff_val < 0:
                    cell.fill = degrade_fill
            except ValueError:
                pass
        elif result['diff'] == 'n/a':
            cell.fill = na_fill

    # 调整列宽
    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12

    # 保存
    wb.save(output_file)
    print(f"\n对比结果已保存: {output_file}")

    # 打印表格（verbose 模式）
    print_results_table(results, verbose)

    # 统计信息
    improved = sum(1 for r in results if r['matched'] and r['diff'] != 'n/a' and float(r['diff']) > 0)
    degraded = sum(1 for r in results if r['matched'] and r['diff'] != 'n/a' and float(r['diff']) < 0)
    unchanged = sum(1 for r in results if r['matched'] and r['diff'] != 'n/a' and float(r['diff']) == 0)
    na_count = sum(1 for r in results if r['diff'] == 'n/a')

    print(f"\n统计:")
    print(f"  性能提升 (+): {improved}")
    print(f"  性能劣化 (-): {degraded}")
    print(f"  无变化 (0):   {unchanged}")
    print(f"  无法对比:     {na_count}")

    return True


def run_cmp_excel(args) -> int:
    """执行 cmp-excel 命令"""
    before_file = args.before
    after_file = args.after
    output_file = args.output
    verbose = args.verbose if hasattr(args, 'verbose') else False

    # 验证文件存在
    if not Path(before_file).exists():
        print(f"错误: 文件不存在: {before_file}", file=sys.stderr)
        return 1
    if not Path(after_file).exists():
        print(f"错误: 文件不存在: {after_file}", file=sys.stderr)
        return 1

    success = compare_ratios(before_file, after_file, output_file, verbose)
    return 0 if success else 1