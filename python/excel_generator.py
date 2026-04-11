#!/usr/bin/env python3
import argparse
import json
import math
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import List, Dict
from collections import defaultdict


def load_compare_result(result_file: str) -> Dict:
    with open(result_file, 'r', encoding='utf-8') as f:
        return json.load(f)


def geometric_mean(values: List[float]) -> float:
    """计算几何平均"""
    if not values or any(v <= 0 for v in values):
        return 0.0
    return math.exp(sum(math.log(v) for v in values) / len(values))


def get_top_level_benchmark(benchmark_name: str) -> str:
    """提取一级 benchmark 名称"""
    if '.' in benchmark_name:
        return benchmark_name.split('.')[0]
    return benchmark_name


def calculate_geometric_means(benchmarks: List[Dict]) -> Dict:
    """计算总的和一级 bench 的几何平均"""
    speedups = [b['speedup'] for b in benchmarks if b['speedup'] > 0]

    if not speedups:
        return {
            'total_geometric_mean': 0.0,
            'top_level_means': {}
        }

    total_gm = geometric_mean(speedups)

    top_level_groups = defaultdict(list)
    for bench in benchmarks:
        if bench['speedup'] > 0:
            top_level = get_top_level_benchmark(bench['benchmark'])
            top_level_groups[top_level].append(bench['speedup'])

    top_level_means = {}
    for top_level, values in top_level_groups.items():
        top_level_means[top_level] = geometric_mean(values)

    return {
        'total_geometric_mean': total_gm,
        'top_level_means': top_level_means
    }


def generate_excel(compare_result: Dict, output_path: str) -> None:
    wb = openpyxl.Workbook()

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    title_font = Font(bold=True, size=14, color='000000')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    server1_name = compare_result['server1']
    server2_name = compare_result['server2']
    commit1 = compare_result['commit1']
    commit2 = compare_result['commit2']
    results = compare_result['benchmarks']
    summary = compare_result['summary']

    geometric_stats = calculate_geometric_means(results)

    ws_summary = wb.create_sheet("概览")

    ws_summary['A1'] = "ASV Benchmark对比报告"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:E1')

    ws_summary['A3'] = "Server1:"
    ws_summary['B3'] = server1_name
    ws_summary['A3'].font = Font(bold=True)

    ws_summary['A4'] = "Server2:"
    ws_summary['B4'] = server2_name
    ws_summary['A4'].font = Font(bold=True)

    ws_summary['A5'] = "Commit1:"
    ws_summary['B5'] = commit1[:8]
    ws_summary['A5'].font = Font(bold=True)

    ws_summary['A6'] = "Commit2:"
    ws_summary['B6'] = commit2[:8]
    ws_summary['A6'].font = Font(bold=True)

    ws_summary['A8'] = "Benchmark数量:"
    ws_summary['B8'] = str(len(results))
    ws_summary['A8'].font = Font(bold=True)

    ws_summary['A10'] = "性能统计"
    ws_summary['A10'].font = Font(bold=True, size=12)

    ws_summary['A11'] = f"{server2_name} 更快:"
    ws_summary['B11'] = str(summary['faster'])

    ws_summary['A12'] = f"{server2_name} 更慢:"
    ws_summary['B12'] = str(summary['slower'])

    ws_summary['A13'] = "性能相同:"
    ws_summary['B13'] = str(summary['same'])

    if results:
        speedups = [r['speedup'] for r in results if r['speedup'] > 0]
        if speedups:
            avg_speedup = sum(speedups) / len(speedups)
            max_speedup = max(speedups)
            min_speedup = min(speedups)

            fastest = max(results, key=lambda x: x['speedup'])
            slowest = min(results, key=lambda x: x['speedup'])

            ws_summary['A15'] = "加速比统计"
            ws_summary['A15'].font = Font(bold=True, size=12)

            ws_summary['A16'] = "平均加速比:"
            ws_summary['B16'] = f"{avg_speedup:.2f}x"

            ws_summary['A17'] = "最大加速比:"
            ws_summary['B17'] = f"{max_speedup:.2f}x"

            ws_summary['A18'] = "最小加速比:"
            ws_summary['B18'] = f"{min_speedup:.2f}x"

            ws_summary['A20'] = "最快benchmark:"
            ws_summary['B20'] = fastest['benchmark']
            ws_summary['C20'] = f"({fastest['speedup']:.2f}x)"

            ws_summary['A21'] = "最慢benchmark:"
            ws_summary['B21'] = slowest['benchmark']
            ws_summary['C21'] = f"({slowest['speedup']:.2f}x)"

    ws_summary['A23'] = "几何平均统计"
    ws_summary['A23'].font = Font(bold=True, size=12)

    ws_summary['A24'] = "总几何平均:"
    ws_summary['B24'] = f"{geometric_stats['total_geometric_mean']:.2f}x"

    row = 25
    for top_level, gm in sorted(geometric_stats['top_level_means'].items()):
        ws_summary[f'A{row}'] = f"{top_level} 几何平均:"
        ws_summary[f'B{row}'] = f"{gm:.2f}x"
        row += 1

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 20

    ws_detail = wb.create_sheet("详细对比")

    headers = ['Benchmark', f'{server1_name}时间(s)', f'{server2_name}时间(s)',
               '加速比', '差异(%)', '性能评估']

    for i, header in enumerate(headers, 1):
        cell = ws_detail.cell(1, i, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for i, item in enumerate(results, 2):
        ws_detail.cell(i, 1, item['benchmark'])
        ws_detail.cell(i, 2, f"{item['time1']:.6f}")
        ws_detail.cell(i, 3, f"{item['time2']:.6f}")
        ws_detail.cell(i, 4, f"{item['speedup']:.2f}")
        ws_detail.cell(i, 5, f"{item['diff_percent']:.2f}")

        diff = item['diff_percent']
        if diff > 10:
            assessment = "明显变慢"
            fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        elif diff > 0:
            assessment = "变慢"
            fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        elif diff < -10:
            assessment = "明显变快"
            fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif diff < 0:
            assessment = "变快"
            fill = PatternFill(start_color='E6F7E6', end_color='E6F7E6', fill_type='solid')
        else:
            assessment = "相同"
            fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        ws_detail.cell(i, 6, assessment)

        for col in range(1, 7):
            cell = ws_detail.cell(i, col)
            cell.border = border
            cell.alignment = Alignment(horizontal='left' if col == 1 else 'center', vertical='center')

        for col in range(1, 7):
            ws_detail.cell(i, col).fill = fill

    ws_detail.column_dimensions['A'].width = 40
    ws_detail.column_dimensions['B'].width = 15
    ws_detail.column_dimensions['C'].width = 15
    ws_detail.column_dimensions['D'].width = 10
    ws_detail.column_dimensions['E'].width = 10
    ws_detail.column_dimensions['F'].width = 12

    ws_detail.freeze_panes = 'A2'

    wb.save(output_path)
    print(f"Excel报告已保存到: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='生成Excel对比报告')
    parser.add_argument('--compare-result', required=True, help='对比结果JSON文件（从benchmark_comparator.py生成）')
    parser.add_argument('--output', required=True, help='输出Excel文件')

    args = parser.parse_args()

    compare_result = load_compare_result(args.compare_result)

    if not compare_result or 'benchmarks' not in compare_result:
        print("错误: 未能加载对比结果")
        exit(1)

    generate_excel(compare_result, args.output)


if __name__ == '__main__':
    main()
