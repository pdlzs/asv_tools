#!/usr/bin/env python3
"""
使用 ASV 命令行进行 benchmark 对比

废弃 ASV API 方式，使用 `asv compare -m machine_name commit1 commit2` 命令行方式
"""
import json
import shutil
import subprocess
from pathlib import Path
from typing import Optional


def get_commit_hash_from_results(asv_dir: str) -> str:
    """
    从 ASV 结果目录中获取最新的 commit hash

    Args:
        asv_dir: ASV 项目目录（包含 results 子目录）

    Returns:
        最新结果的 commit hash
    """
    results_dir = Path(asv_dir) / "results"

    if not results_dir.exists():
        raise ValueError(f"结果目录不存在: {results_dir}")

    # 遍历所有机器目录，找到最新的结果
    latest_commit = None
    latest_date = 0

    for machine_dir in results_dir.iterdir():
        if not machine_dir.is_dir():
            continue

        for json_file in machine_dir.glob("*.json"):
            if json_file.name == "machine.json":
                continue

            try:
                with open(json_file, 'r') as f:
                    data = json.load(f)
                    commit_hash = data.get("commit_hash", "")
                    date_val = data.get("date", 0)

                    if commit_hash and date_val > latest_date:
                        latest_date = date_val
                        latest_commit = commit_hash
            except (json.JSONDecodeError, IOError):
                continue

    if not latest_commit:
        raise ValueError(f"在 {asv_dir} 中未找到任何 commit 结果")

    return latest_commit


def get_machine_name(asv_dir: str) -> str:
    """
    从 ASV 结果目录中获取机器名称

    Args:
        asv_dir: ASV 项目目录

    Returns:
        机器名称
    """
    results_dir = Path(asv_dir) / "results"

    for machine_dir in results_dir.iterdir():
        if machine_dir.is_dir():
            machine_json = machine_dir / "machine.json"
            if machine_json.exists():
                try:
                    with open(machine_json, 'r') as f:
                        data = json.load(f)
                        return data.get("machine", machine_dir.name)
                except (json.JSONDecodeError, IOError):
                    return machine_dir.name

    raise ValueError(f"在 {asv_dir} 中未找到机器信息")


def filter_ratio_na_lines(table_output: str, skip_ratio_na: bool = False) -> str:
    """过滤掉 Ratio 为 n/a 的行

    Args:
        table_output: ASV compare 输出的表格文本
        skip_ratio_na: 是否跳过 Ratio 为 n/a 的行

    Returns:
        过滤后的表格文本（如果 skip_ratio_na=True），否则返回原文本
    """
    if not skip_ratio_na:
        return table_output

    lines = [line for line in table_output.split('\n') if line.strip()]

    # 找到表格开始位置
    table_start_idx = -1
    for i, line in enumerate(lines):
        if '|' in line and ('Change' in line or 'Before' in line or 'After' in line):
            table_start_idx = i
            break

    if table_start_idx == -1:
        return table_output

    # 保留表格之前的内容（如 "All benchmarks" 等说明）
    prefix_lines = lines[:table_start_idx]
    table_lines = lines[table_start_idx:]

    # 找到 Ratio 列的索引
    ratio_col_idx = -1
    if table_lines:
        header_cols = [col.strip() for col in table_lines[0].split('|')]
        header_cols = header_cols[1:-1]  # 移除首尾空列
        for idx, col in enumerate(header_cols):
            if col == 'Ratio':
                ratio_col_idx = idx
                break

    if ratio_col_idx < 0:
        return table_output  # 没有 Ratio 列，不过滤

    # 过滤掉 Ratio 为 n/a 的行
    filtered_lines = []
    filtered_count = 0
    for i, line in enumerate(table_lines):
        if i <= 1:  # 保留标题行和分隔行
            filtered_lines.append(line)
            continue

        cols = [col.strip() for col in line.split('|')]
        cols = cols[1:-1]  # 移除首尾空列

        if len(cols) > ratio_col_idx:
            ratio_val = cols[ratio_col_idx]
            if ratio_val == 'n/a':
                filtered_count += 1
                continue  # 跳过 Ratio 为 n/a 的行

        filtered_lines.append(line)

    if filtered_count > 0:
        print(f"已过滤 {filtered_count} 行 Ratio 为 n/a 的数据")

    # 合并前缀和过滤后的表格
    result_lines = prefix_lines + filtered_lines
    return '\n'.join(result_lines)


def generate_excel_from_table(
    table_output: str,
    output_dir: Path,
    server1: str,
    server2: str,
    timestamp: Optional[str] = None
):
    """从 ASV 表格输出生成 Excel 文件

    Args:
        table_output: ASV compare 输出的表格文本（已过滤）
        output_dir: 输出目录
        server1: 服务器1名称
        server2: 服务器2名称
        timestamp: 时间戳（可选）
    """
    try:
        import openpyxl
        import openpyxl.utils
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("警告: 未安装 openpyxl，跳过 Excel 表格生成")
        return

    lines = [line for line in table_output.split('\n') if line.strip()]

    # 找到表格开始位置
    table_start_idx = -1
    for i, line in enumerate(lines):
        if '|' in line and ('Change' in line or 'Before' in line or 'After' in line):
            table_start_idx = i
            break

    if table_start_idx == -1:
        print("警告: 无法解析表格，跳过 Excel 生成")
        return

    table_lines = lines[table_start_idx:]

    # 文件名：带时间戳（与输出目录同名）
    file_suffix = f"_{timestamp}" if timestamp else ""
    excel_output_path = output_dir / f"{server1}_vs_{server2}_table{file_suffix}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ASV Compare Table"

    for i, line in enumerate(table_lines):
        if not line.strip():
            continue

        cols = [col.strip() for col in line.split('|')]
        cols = cols[1:-1]  # 移除首尾空列

        for j, col in enumerate(cols):
            cell = ws.cell(row=i+1, column=j+1, value=col)

            if i == 0:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            elif i == 1:
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # 自动调整列宽
    for col_idx, column in enumerate(ws.columns, start=1):
        max_length = 0
        column_cells = [cell for cell in column]
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        adjusted_width = min(adjusted_width, 50)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

    wb.save(excel_output_path)
    print(f"表格已保存到: {excel_output_path}")


def modify_commit_hash(json_file: Path, server_name: str, target_machine: str) -> str:
    """
    修改结果文件的 commit_hash 和 machine 名称

    Args:
        json_file: 结果 json 文件路径
        server_name: 服务器名称（用于 commit_hash 前缀）
        target_machine: 目标机器名称（用于 params.machine）

    Returns:
        修改后的 commit hash
    """
    with open(json_file, 'r') as f:
        data = json.load(f)

    original_commit = data.get("commit_hash", "")
    if not original_commit:
        raise ValueError(f"{json_file} 中没有 commit_hash 字段")

    # 添加服务器前缀（使用 - 分隔符，ASV 显示前8位时更友好）
    modified_commit = f"{server_name}-{original_commit}"
    data["commit_hash"] = modified_commit

    # 修改 params.machine 为目标机器名称
    if "params" in data and "machine" in data["params"]:
        data["params"]["machine"] = target_machine

    # 写回文件
    with open(json_file, 'w') as f:
        json.dump(data, f)

    return modified_commit


def compare_results(
    asv_dir1: str,
    asv_dir2: str,
    server1_name: str,
    server2_name: str,
    output_dir: Path,
    show_all: bool = True,
    verbose: bool = False,
    timestamp: Optional[str] = None,
    skip_excel: bool = False,
    skip_ratio_na: bool = False
) -> bool:
    """
    执行 ASV 结果对比（使用命令行方式）

    Args:
        asv_dir1: 服务器1的 ASV 目录
        asv_dir2: 服务器2的 ASV 目录
        server1_name: 服务器1名称
        server2_name: 服务器2名称
        output_dir: 输出目录
        show_all: 是否显示所有 benchmark
        verbose: 是否显示详细输出
        timestamp: 时间戳（可选）
        skip_excel: 是否跳过生成 Excel 文件
        skip_ratio_na: 是否跳过 Ratio 为 n/a 的行（影响 TXT 和 Excel）

    Returns:
        成功返回 True
    """
    # 获取原始 commit hash
    original_commit1 = get_commit_hash_from_results(asv_dir1)
    original_commit2 = get_commit_hash_from_results(asv_dir2)

    print(f"Server1 ({server1_name}) original commit: {original_commit1}")
    print(f"Server2 ({server2_name}) original commit: {original_commit2}")

    # 获取机器名称
    machine1 = get_machine_name(asv_dir1)
    machine2 = get_machine_name(asv_dir2)

    # 使用统一的目标机器名称（用于 ASV compare）
    # 使用一个通用的名称，避免因机器名不同导致问题
    target_machine = "compare_machine"

    print(f"Server1 机器名称: {machine1}")
    print(f"Server2 机器名称: {machine2}")
    print(f"使用统一机器名称: {target_machine}")

    # 创建合并的 ASV 目录
    merged_dir = output_dir / "merged_asv"
    merged_results_base = merged_dir / "results"  # results 目录（不含机器名）
    merged_results_dir = merged_results_base / target_machine  # results/<machine> 目录

    merged_dir.mkdir(parents=True, exist_ok=True)
    merged_results_base.mkdir(parents=True, exist_ok=True)
    merged_results_dir.mkdir(parents=True, exist_ok=True)

    # 复制 benchmarks.json 到 results 目录（ASV 期望在这里找到它）
    benchmarks1 = Path(asv_dir1) / "results" / "benchmarks.json"
    benchmarks2 = Path(asv_dir2) / "results" / "benchmarks.json"

    if benchmarks1.exists():
        shutil.copy2(benchmarks1, merged_results_base / "benchmarks.json")
    elif benchmarks2.exists():
        shutil.copy2(benchmarks2, merged_results_base / "benchmarks.json")

    # 复制两个服务器的结果文件到合并目录
    # 注意：两台服务器可能有不同的机器名称，分别从各自的目录复制
    results1_dir = Path(asv_dir1) / "results" / machine1
    results2_dir = Path(asv_dir2) / "results" / machine2

    commit1 = None
    commit2 = None

    def copy_and_modify_results(results_dir: Path, server_name: str, original_commit: str, source_machine: str, target_machine: str) -> str:
        """复制结果文件并修改 commit hash 和 machine 名称，返回修改后的 commit hash"""
        modified_commit = None

        if not results_dir.exists():
            print(f"警告: 结果目录不存在: {results_dir}")
            return None

        for f in results_dir.glob("*.json"):
            if f.name == "machine.json":
                # 修改 machine.json 中的机器名称为目标名称
                with open(f, 'r') as rf:
                    machine_data = json.load(rf)
                machine_data["machine"] = target_machine
                with open(merged_results_dir / "machine.json", 'w') as wf:
                    json.dump(machine_data, wf)
            else:
                # 构建新文件名：使用新的 commit hash 前缀
                new_prefix = f"{server_name}-{original_commit[:8]}"
                new_filename = f"{new_prefix}-{f.name.split('-', 1)[1]}" if "-" in f.name else f"{new_prefix}"
                target_path = merged_results_dir / new_filename

                # 复制并修改 commit_hash 和 params.machine
                shutil.copy2(f, target_path)
                modified_commit = modify_commit_hash(target_path, server_name, target_machine)

                if verbose:
                    print(f"  {server_name} ({source_machine}): {f.name} -> {new_filename}")

        return modified_commit

    # 复制 server1 的结果
    if results1_dir.exists():
        commit1 = copy_and_modify_results(results1_dir, server1_name, original_commit1, machine1, target_machine)
    else:
        print(f"错误: server1 结果目录不存在: {results1_dir}")

    # 复制 server2 的结果
    if results2_dir.exists():
        commit2 = copy_and_modify_results(results2_dir, server2_name, original_commit2, machine2, target_machine)
    else:
        print(f"错误: server2 结果目录不存在: {results2_dir}")

    if not commit1 or not commit2:
        print("错误: 无法获取 commit hash")
        return False

    print(f"\n使用 commit: {commit1[:8]} ({server1_name}) vs {commit2[:8]} ({server2_name})")

    # 创建 asv.conf.json
    asv_conf = {
        "version": 1,
        "project": "benchmark",
        "repo": ".",
        "results_dir": "results",
        "benchmark_dir": "benchmarks",
        "dvcs": "git"
    }
    with open(merged_dir / "asv.conf.json", 'w') as f:
        json.dump(asv_conf, f, indent=2)

    # 创建一个简单的 git repo（ASV compare 需要）
    # ASV 会尝试克隆 repo，但只是为了获取 commit 信息
    # 我们创建一个空 repo 并添加虚假的 commit
    if not (merged_dir / ".git").exists():
        subprocess.run(["git", "init"], cwd=str(merged_dir), capture_output=True)
        subprocess.run(["git", "config", "user.email", "asv@test.com"], cwd=str(merged_dir), capture_output=True)
        subprocess.run(["git", "config", "user.name", "ASV Test"], cwd=str(merged_dir), capture_output=True)

        # 创建一个空的 benchmark 目录
        benchmark_dir = merged_dir / "benchmarks"
        benchmark_dir.mkdir(exist_ok=True)

        # 创建一个空的 __init__.py 文件
        (benchmark_dir / "__init__.py").touch()

        # 添加并提交
        subprocess.run(["git", "add", "."], cwd=str(merged_dir), capture_output=True)
        subprocess.run(["git", "commit", "-m", "Initial commit"], cwd=str(merged_dir), capture_output=True)

        # 创建虚假的 commit 来匹配我们的 commit hash 前缀
        # 由于 commit hash 是服务器前缀 + 原始 hash，我们无法真正创建匹配的 commit
        # 但 ASV compare 只需要 repo 存在，commit hash 用于查找结果文件

    # 构建 asv compare 命令
    cmd = [
        "asv", "compare",
        "-m", target_machine,
        "--factor", "1.0",
        "--sort", "default",
    ]
    if not show_all:
        cmd.append("--only-changed")
    cmd.extend([commit1, commit2])

    print(f"\n执行: {' '.join(cmd)}")

    # 执行 asv compare 命令
    result = subprocess.run(
        cmd,
        cwd=str(merged_dir),
        capture_output=True,
        text=True
    )

    table_output = result.stdout
    if result.stderr:
        print(f"stderr: {result.stderr}")

    if result.returncode != 0:
        print(f"错误: asv compare 返回码 {result.returncode}")
        print(f"stdout: {result.stdout}")
        return False

    print(table_output)

    # 过滤 Ratio 为 n/a 的行（如果启用）
    filtered_table_output = filter_ratio_na_lines(table_output, skip_ratio_na)

    # 文件名：带时间戳（与输出目录同名）
    file_suffix = f"_{timestamp}" if timestamp else ""
    txt_output_path = output_dir / f"{server1_name}_vs_{server2_name}_table{file_suffix}.txt"
    with open(txt_output_path, 'w', encoding='utf-8') as f:
        header = f"Compare: {server1_name} [{commit1}] vs {server2_name} [{commit2}]\n\n"
        f.write(header + filtered_table_output)
    print(f"\n表格已保存到: {txt_output_path}")

    # 生成 Excel 文件
    if not skip_excel:
        generate_excel_from_table(filtered_table_output, output_dir, server1_name, server2_name, timestamp)

    return True